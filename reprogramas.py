import streamlit as st
import requests
import zipfile
import io
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
import plotly.express as px
import plotly.graph_objects as go
import openpyxl

# --- 1. CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Supervisión YUPANA - Osinergmin", layout="wide", initial_sidebar_state="expanded")
st.title("⚡ Dashboard de Supervisión Continua - Programas y Reprogramas")
st.markdown("Fiscalización Dinámica Multi-Día de Curvas de Carga, Reprogramaciones y Capacidad Inactiva")

# --- 2. PARÁMETROS OPERATIVOS ---
MES_TXT = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]
barras = ["SANTA ROSA 220 A", "MOQUEGUA 220", "ZORRITOS 220"]
rdo_letras = list("ABCDEFGHIJ")

archivos_clave = {
    "HIDRO"   : "Hidro - Despacho (MW)",
    "TERMICA" : "Termica - Despacho (MW)",
    "RER"     : "Rer y No COES - Despacho (MW)",
    "CMG"     : "CMg - Barra ($ por MWh)",
    "POT_EFEC": "Termica - Potencia Efectiva (MW)"
}

# --- MAPA CROMÁTICO Y DE APILAMIENTO OSINERGMIN ---
COLOR_MAP = {
    "Biogás+Biomasa+Nafta+Flexigas": "#800080",    # Morado
    "Solar": "#FFD700",                            # Amarillo
    "Eólica": "#808080",                           # Plomo
    "Hidráulica": "#1f77b4",                       # Azul
    "Gas de Camisea": "#006400",                   # Verde oscuro
    "Gas del Norte+Gas de la Selva": "#90EE90",    # Verde claro
    "Residual+Diésel D2": "#FF0000"                # Rojo
}

ORDEN_TECNOLOGIAS = [
    "Biogás+Biomasa+Nafta+Flexigas",
    "Solar",
    "Eólica",
    "Hidráulica",
    "Gas de Camisea",
    "Gas del Norte+Gas de la Selva",
    "Residual+Diésel D2"
]

# --- 3. CLASIFICADOR TERMODINÁMICO MAESTRO ---
def clasificar_tecnologia_yupana(nombre_central, origen_archivo=""):
    nombre = str(nombre_central).upper()
    
    biomasa_kws = [
        "PARAMONGA", "JACINTO", "HUAYCOLORO", "GRINGA", "MAPLE", "FLEXIGAS", "NAFTA", 
        "LUREN", "BIOMASA", "BIOGAS", "AGROAURORA", "CAHUAPANAS", "SUPE", "LAREDO", 
        "PETRAMAS", "DOÑA CATALINA", "DONA CATALINA", "PORTILLO", "CASA GRANDE", "CASAGRANDE",
        "CANA BRAVA", "AGROOLMOS", "CALLAO", "REFTALARA"
    ]
    if any(kw in nombre for kw in biomasa_kws): return "Biogás+Biomasa+Nafta+Flexigas"
        
    solar_kws = [
        "SOL", "PANAMERICANA", "RUBI", "INTIPAMPA", "CLEMESI", "MATARANI", 
        "REPARTICION", "MAJES", "MISTI", "TACNA", "CS CARHUAQUERO", 
        "CSCOENERGY", "CSF", "CSSUNNY", "CS SAN MARTIN", "CSSANMARTIN"
    ]
    if any(kw in nombre for kw in solar_kws): return "Solar"
        
    eolica_kws = ["EOL", "WAYRA", "LOMITAS", "CUPISNIQUE", "PE TALARA", "MARCONA", "TRES HERMANAS", "DUNA", "HUAMBOS", "SAN JUAN"]
    if any(kw in nombre for kw in eolica_kws): return "Eólica"
        
    hidro_kws = ["HIDRO", "CH ", "C.H.", "MANTARO", "RESTITUCION", "CHAGLLA", "CERRO DEL AGUILA", "MACHUPICCHU", "HUINCO", "CHARCANI", "CAÑON DEL PATO", "SAN GABAN", "CHIMAY", "PLATANAL", "YUNCHAN", "QUISHUAR", "AURA", "ZONGO", "CARPAPATA", "LA JOYA", "STACRUZ", "HUASAHUASI", "RONCADOR", "PURMACANA", "NIMPERIAL", "PIZARRAS", "POECHOS", "CANCHAYLLO", "CHANCAY", "RUCUY", "RUNATULLO", "YANAPAMPA", "POTRERO", "YARUCAYA", "CHANGELI", "8AGOSTO", "RENOVANDESH", "EL CARMEN", "TUPURI", "HUALLIN", "GALLITO", "YAUPI", "MATUCANA", "CALLAHUANCA", "MOYOPAMPA", "HUANZA", "CHEO", "CHURO", "CHHER", "CHZANA", "CURUMUY", "PIAS"]
    if origen_archivo == "HIDRO" or any(kw in nombre for kw in hidro_kws): return "Hidráulica"
        
    # 1ro: EVALUAR SI ES DIÉSEL EXPLICITAMENTE (Atrapa casos como "FENIX CCOMB GT12 D2")
    diesel_kws = ["D2", "R6", "RESIDUAL", "DIESEL", "ILO21", "ILO 21", "ILO1", "ILO 1", "MOLLENDO", "RECKA", "INDEPENDENCIA", "SAMANCO", "TARAPOTO", "IQUITOS", "YURIMAGUAS", "PUERTO MALDONADO", "BELLAVISTA", "PEDRO RUIZ", "ETEN", "PIURA D", "CALANA", "ELOR", "SHCUMMINS", "SNTV", "NEPI", "PUERTO BRAVO", "NODO"]
    if any(kw in nombre for kw in diesel_kws): return "Residual+Diésel D2"

    # 2do: FILTRO DE CICLOS COMBINADOS Y DUALES (Si llegó aquí, es porque NO tiene sufijo D2/Residual explícito, por tanto es Gas)
    duales_gas_kws = ["FENIX", "KALLPA", "CHILCA", "VENTANILLA", "LAS FLORES", "SANTO DOMINGO", "MALACAS", "TALLANCA", "AGUAYTIA", "TERMOSELVA"]
    if any(ex in nombre for ex in duales_gas_kws):
        if any(kw in nombre for kw in ["MALACAS", "TALLANCA", "AGUAYTIA", "TERMOSELVA"]):
            return "Gas del Norte+Gas de la Selva"
        return "Gas de Camisea"
            
    gas_norte_kws = ["AGUAYTIA", "TERMOSELVA", "PUCALLPA", "MALACAS", "ZORRITOS", "PARIÑAS", "EEEP", "ENEL PIURA", "PIURA G", "NUEVA ZORRITOS", "AGE", "TALLANCA", "MAL2", "TABLAZO"]
    if any(kw in nombre for kw in gas_norte_kws): return "Gas del Norte+Gas de la Selva"
        
    return "Gas de Camisea"

# --- 4. INGESTA Y ETL REDUNDANTE ---
def cargar_df_desde_zip(zf, stem):
    for info in zf.infolist():
        nombre_base = info.filename.split('/')[-1]
        if stem in nombre_base and not nombre_base.startswith("~"):
            with zf.open(info) as f:
                if nombre_base.upper().endswith('.CSV'): 
                    try: df = pd.read_csv(f, sep=None, engine='python')
                    except:
                        f.seek(0)
                        df = pd.read_csv(f, sep=',')
                    return df
                elif nombre_base.upper().endswith(('.XLSX', '.XLS')): 
                    return pd.read_excel(f, engine='openpyxl')
    return None

def extraer_todas_centrales(df):
    series = {}
    if df is None or df.empty: return series
    invalid_cols = ["HORA", "TIEMPO", "FECHA", "ETAPA", "GENERADOR"]
    if df.shape[1] > 1:
        cols = [c for c in df.columns if not any(inv in str(c).upper() for inv in invalid_cols) and not str(c).startswith("Unnamed")]
        for c in cols: series[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).tolist()
    else:
        enc = [h.strip() for h in str(df.columns[0]).split(",")]
        start_idx = 0
        if len(enc) < 2:
            enc = [h.strip() for h in str(df.iloc[0,0]).split(",")]
            start_idx = 1
        nombres_validos, idx_validos = [], []
        for i, nombre in enumerate(enc[1:], start=1):
            if not any(inv in nombre.upper() for inv in invalid_cols):
                nombres_validos.append(nombre)
                idx_validos.append(i)
                series[nombre] = []
        for fila in df.iloc[start_idx:, 0].astype(str):
            partes = [p.strip() for p in fila.split(",")]
            for nombre, i in zip(nombres_validos, idx_validos):
                series[nombre].append(float(partes[i]) if i < len(partes) and partes[i] else 0.0)
    return series

def extraer_columna(df, col):
    return pd.to_numeric(df[col], errors='coerce').fillna(0).tolist() if df is not None and col in df.columns else None

def rellenar_hasta_48(lst):
    if not lst: return [0.0]*48
    faltan = 48 - len(lst)
    return ([0.0]*faltan + lst) if faltan > 0 else lst[:48]

def suma_elementos_variable(*listas):
    if not listas: return []
    length = max(len(l) for l in listas if l)
    if length == 0: return []
    out = [0.0]*length
    for lst in listas:
        if lst:
            for i in range(min(length, len(lst))):
                if pd.notna(lst[i]): out[i] += lst[i]
    return out

def renombrar_con_sufijos(diccionario_series, tipo_origen):
    renamed = {}
    for c, vals in diccionario_series.items():
        c_clean = str(c).replace("(EOL)", "").replace("(SOL)", "").replace("(HID)", "").replace("(TER)", "").replace("(RER)", "").strip()
        cat_maestra = clasificar_tecnologia_yupana(c_clean, tipo_origen)
        if cat_maestra == "Hidráulica": renamed[f"{c_clean} (HID)"] = vals
        elif cat_maestra == "Eólica": renamed[f"{c_clean} (EOL)"] = vals
        elif cat_maestra == "Solar": renamed[f"{c_clean} (SOL)"] = vals
        else: renamed[f"{c_clean} (TER)"] = vals
    return renamed

# EXTRAE EL MOTIVO DINÁMICAMENTE BUSCANDO LA PALABRA "MOTIVO" EN LA COLUMNA C
def extraer_motivo_dinamico(y, m, M, d, ddmm, l, headers):
    urls = [
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx",
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}%20{l}%2FReprog_{ddmm}{l}.xlsx",
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2F{ddmm}{l}.xlsx"
    ]
    for u in urls:
        try:
            r = requests.get(u, headers=headers, timeout=10)
            if r.status_code == 200 and len(r.content) > 1000:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
                ws = wb.worksheets[0]
                
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=3).value
                    if cell_value and "MOTIVO" in str(cell_value).upper():
                        motivo_val = ws.cell(row=row+1, column=4).value
                        if motivo_val:
                            return str(motivo_val).strip()
                        else:
                            return "Motivo encontrado pero vacío en la celda contigua."
                            
                return "No se encontró la palabra 'MOTIVO' en la columna C."
        except:
            pass
    return "No se pudo extraer el archivo de origen o hubo un error de lectura."

@st.cache_data(show_spinner=False)
def extraer_datos_dia_memoria(f):
    y, m, d = f.strftime("%Y"), f.strftime("%m"), f.strftime("%d")
    M = MES_TXT[f.month-1]
    fecha_str = f"{y}{m}{d}"
    ddmm = f"{d}{m}"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
    }
    
    urls_a_intentar = {
        "PDO": [f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FYUPANA_{fecha_str}.zip"]
    }
    
    for l in rdo_letras:
        urls_a_intentar[f"RDO_{l}"] = [
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FYUPANA_{ddmm}{l}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}%20{l}%2FYUPANA_{ddmm}{l}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FYUPANA_{fecha_str}{l}.zip"
        ]

    datos_dia = {"Dataframes": {}, "Log": []}
    for nombre, lista_enlaces in urls_a_intentar.items():
        exito = False
        for enlace in lista_enlaces:
            if exito: break
            try:
                r = requests.get(enlace, headers=headers, timeout=15)
                if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
                    with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                        datos_dia["Dataframes"][nombre] = {}
                        for key, stem in archivos_clave.items():
                            datos_dia["Dataframes"][nombre][key] = cargar_df_desde_zip(zf, stem)
                    
                    if "RDO" in nombre:
                        letra = nombre.split("_")[-1]
                        motivo = extraer_motivo_dinamico(y, m, M, d, ddmm, letra, headers)
                        datos_dia["Dataframes"][f"MOTIVO_{nombre}"] = motivo
                        
                    datos_dia["Log"].append(f"✅ {nombre}")
                    exito = True
            except Exception:
                continue 
        
        if not exito:
            if f == date.today() and "RDO" in nombre:
                datos_dia["Log"].append(f"⏳ {nombre} (Aún no emitido)")
            else:
                datos_dia["Log"].append(f"❌ {nombre} (No publicado)")
                
    return datos_dia

# --- 5. MOTOR GRÁFICO MAESTRO MULTIDÍA (SIN TÍTULOS NATIVOS) ---
def crear_grafica_area_apilada(df_plot, marcadores=None, aplicar_colores=False, orden_fijo=None):
    df_plot = df_plot.fillna(0)
    num_cols = [c for c in df_plot.columns if c != 'Hora']
    df_plot[num_cols] = df_plot[num_cols].apply(pd.to_numeric, errors='coerce').fillna(0).round(2)
    
    df_plot['TOTAL_SISTEMA'] = df_plot[num_cols].sum(axis=1).round(2)
    idx_pico = df_plot['TOTAL_SISTEMA'].idxmax()
    pico_mw = df_plot.loc[idx_pico, 'TOTAL_SISTEMA']
    pico_hora = df_plot.loc[idx_pico, 'Hora']
    
    if orden_fijo:
        orden_columnas = [col for col in orden_fijo if col in df_plot.columns]
    else:
        totales_por_unidad = df_plot.drop(columns=['Hora', 'TOTAL_SISTEMA']).sum()
        orden_columnas = totales_por_unidad.sort_values(ascending=False).index.tolist()
    
    cols_mantener = ['Hora', 'TOTAL_SISTEMA'] + orden_columnas
    df_melt = df_plot[cols_mantener].melt(id_vars=['Hora', 'TOTAL_SISTEMA'], var_name='Unidad Generadora', value_name='Potencia_MW')
    
    # Se conserva la base matemática para evitar polígonos rotos
    df_melt['Potencia_Plot'] = df_melt['Potencia_MW']
    
    kw_args = {"data_frame": df_melt, "x": "Hora", "y": "Potencia_Plot", "color": "Unidad Generadora", "labels": {"Potencia_Plot": "Potencia Activa (MW)"}}
    if aplicar_colores: kw_args["color_discrete_map"] = COLOR_MAP
    
    fig = px.area(**kw_args)
    fig.update_xaxes(tickformat="%d/%m %H:%M", tickangle=45)
    
    fig.add_scatter(x=df_plot['Hora'], y=df_plot['TOTAL_SISTEMA'], mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>⚡ TOTAL SISTEMA</b>', showlegend=False)
    
    # HOVER INTELIGENTE: Ignora 0 MW y agrega la fecha/hora debajo de la potencia
    for trace in fig.data:
        y_vals = trace.y
        hover_flags = []
        for val in y_vals:
            try:
                if pd.isna(val) or float(val) <= 0.01:
                    hover_flags.append('skip')
                else:
                    hover_flags.append('all')
            except:
                hover_flags.append('all')
        
        trace.hoverinfo = hover_flags
        
        if 'TOTAL SISTEMA' in trace.name:
            trace.hovertemplate = '<b>%{y:,.2f} MW</b><br>%{x|%d/%m %H:%M}'
        else:
            trace.hovertemplate = "%{y:,.2f} MW"
    
    if marcadores:
        for ts, texto in marcadores:
            fig.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(255,255,255,0.7)")
            
            texto_limpio = texto.replace("(", "").replace(")", "")
            texto_con_hora = f"{texto_limpio} {ts.strftime('%H:%M')}"
            align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
            
            fig.add_annotation(
                x=ts, y=1.02, yref="paper", text=f"<b>{texto_con_hora}</b>", showarrow=False, 
                font=dict(size=10, color="white"), bgcolor="#e74c3c", bordercolor="white", 
                borderwidth=1, borderpad=3, textangle=-90, yanchor="bottom", xanchor=align
            )
            
    fig.add_annotation(x=pico_hora, y=pico_mw, text=f"<b>Pico Máximo: {pico_mw:,.2f} MW</b><br>{pico_hora.strftime('%d/%m %H:%M')}", showarrow=True, arrowhead=2, arrowsize=1.5, arrowwidth=2, arrowcolor="#e74c3c", ax=0, ay=-50, font=dict(size=12, color="#c0392b"), bgcolor="rgba(255,255,255,0.8)", bordercolor="#c0392b", borderwidth=1, borderpad=4)
    
    fig.update_layout(hovermode="x unified", height=650, margin=dict(t=150, b=50, l=60, r=50))
    return fig

# --- 6. INTERFAZ Y EJECUCIÓN ---
st.sidebar.header("Parámetros de Fiscalización")
rango_fechas = st.sidebar.date_input("Intervalo de Fechas Continuo", value=(date.today() - timedelta(days=1), date.today()))

if st.sidebar.button("Construir Matriz de Operación Continua", type="primary"):
    if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
        ini, fin = rango_fechas
        st.session_state['fecha_ini'], st.session_state['fecha_fin'] = ini, fin
        
        status, prog_bar = st.empty(), st.progress(0)
        log_exp = st.expander("Ver bitácora de descargas (COES)", expanded=False)
        
        datos_completos = {}
        dias = (fin - ini).days + 1
        for k in range(dias):
            f_actual = ini + timedelta(days=k)
            status.markdown(f"**⏳ Vectorizando Días (In-Memory):** {f_actual.strftime('%d/%m/%Y')} *(Día {k+1}/{dias})*")
            datos_dia = extraer_datos_dia_memoria(f_actual)
            datos_completos[f_actual] = datos_dia
            with log_exp: st.markdown(f"**{f_actual.strftime('%d/%m/%Y')}** ➔ " + " | ".join(datos_dia["Log"]))
            prog_bar.progress((k + 1) / dias)
            
        st.session_state['datos_yupana'] = datos_completos
        status.success("✅ Motor Dimensional Compilado con Éxito.")
        prog_bar.empty()

# --- 7. VISUALIZACIÓN DINÁMICA MULTIDÍA ---
if 'datos_yupana' in st.session_state:
    data = st.session_state['datos_yupana']
    fechas_ordenadas = sorted(data.keys())
    
    active_prog_dict = {}
    ts_dict = {}
    dics_cache_dict = {}
    marcadores_globales = []
    timestamps_globales = []
    
    for f in fechas_ordenadas:
        df_dia_sel = data[f]["Dataframes"]
        progs = [p for p in ["PDO"] + [f"RDO_{l}" for l in rdo_letras] if p in df_dia_sel]
        if not progs: continue
        
        dics_cache = {}
        for p in progs:
            dics_cache[p] = {
                "HIDRO": extraer_todas_centrales(df_dia_sel[p].get("HIDRO")),
                "TERMICA": extraer_todas_centrales(df_dia_sel[p].get("TERMICA")),
                "RER": extraer_todas_centrales(df_dia_sel[p].get("RER")),
                "POT_EFEC": extraer_todas_centrales(df_dia_sel[p].get("POT_EFEC")),
                "CMG": df_dia_sel[p].get("CMG")
            }
            
        active_prog = [progs[0]] * 48
        if len(progs) > 1:
            for p in progs[1:]:
                tot = [0.0]*48
                for v in dics_cache[p]["HIDRO"].values(): tot = suma_elementos_variable(tot, rellenar_hasta_48(v))
                for v in dics_cache[p]["TERMICA"].values(): tot = suma_elementos_variable(tot, rellenar_hasta_48(v))
                for v in dics_cache[p]["RER"].values(): tot = suma_elementos_variable(tot, rellenar_hasta_48(v))
                for i, val in enumerate(tot):
                    if val > 100: 
                        for j in range(i, 48): active_prog[j] = p
                        break
                        
        ts_dia = [datetime.combine(f, datetime.min.time()) + timedelta(minutes=30*(i+1)) for i in range(48)]
        p_actual = active_prog[0]
        marcadores_globales.append((ts_dia[0], p_actual))
        for i in range(1, 48):
            if active_prog[i] != p_actual:
                p_actual = active_prog[i]
                marcadores_globales.append((ts_dia[i], p_actual))
                
        active_prog_dict[f] = active_prog
        ts_dict[f] = ts_dia
        dics_cache_dict[f] = dics_cache
        timestamps_globales.extend(ts_dia)

    t_cmg, t_hidro, t_term, t_inactiva, t_dem, t_eol, t_sol, t_motivos_rdo = st.tabs([
        "💸 CMG", "💧 Despacho Hidro", "🔥 Despacho Térmico", 
        "🛑 Inactiva Diésel", "📈 Demanda y Generación", "💨 Eólico", "☀️ Solar", "📋 Motivos RDO"
    ])

    # === CMG (SIN LÍMITES / RAW DATA) ===
    with t_cmg:
        st.markdown("### 💸 Evolución Continua del CMG")
        dfs_cmg = []
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            dia_cmg = {b: [0.0]*48 for b in barras}
            for i in range(48):
                p = active_prog[i]
                df_c = dics_cache[p]["CMG"]
                
                if df_c is not None:
                    df_c_cols = {str(col).strip(): col for col in df_c.columns}
                    for b in barras:
                        b_strip = b.strip()
                        if b_strip in df_c_cols:
                            col_real = df_c_cols[b_strip]
                            v_lst = rellenar_hasta_48(extraer_columna(df_c, col_real))
                            dia_cmg[b][i] = v_lst[i]
                            
            df_dia = pd.DataFrame(dia_cmg)
            df_dia.insert(0, 'Hora', ts_dict[f])
            dfs_cmg.append(df_dia)
            
        if dfs_cmg:
            df_cmg_plot = pd.concat(dfs_cmg, ignore_index=True)
            fig_cmg = px.line(df_cmg_plot, x='Hora', y=barras, labels={"value": "USD/MWh", "variable": "Barra"})
            
            for trace in fig_cmg.data:
                trace.hovertemplate = '<b>%{y:,.2f} USD/MWh</b><br>%{x|%d/%m %H:%M}'
            
            for ts, txt in marcadores_globales:
                fig_cmg.add_vline(x=ts, line_width=1, line_dash="dash", line_color="grey")
                txt_limpio = txt.replace("(", "").replace(")", "")
                txt_final = f"{txt_limpio} {ts.strftime('%d/%m %H:%M')}"
                align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                fig_cmg.add_annotation(x=ts, y=1.02, yref="paper", text=f"<b>{txt_final}</b>", showarrow=False, font=dict(size=10, color="black"), bgcolor="lightgrey", textangle=-90, yanchor="bottom", xanchor=align)
                
            fig_cmg.update_layout(hovermode="x unified", height=550, margin=dict(t=150, l=60))
            st.plotly_chart(fig_cmg, use_container_width=True)

    # === RUTINA MAESTRA (EXTRAE ESTRICTAMENTE DE MATRIZ ACTIVA: HIDRO, TERMICA, RER) ===
    def render_tab_generico(tipo_principal):
        dfs_tab = []
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            dia_data = {}
            for i in range(48):
                p = active_prog[i]
                for arch in ["HIDRO", "TERMICA", "RER"]:
                    if not dics_cache[p][arch]: continue
                    
                    dic_renombrado = renombrar_con_sufijos(dics_cache[p][arch], arch if arch in ["HIDRO", "RER"] else "TERMICA")
                    
                    for k, v_list in dic_renombrado.items():
                        c_clean = str(k).replace(" (TER)", "").replace(" (HID)", "").replace(" (SOL)", "").replace(" (EOL)", "").strip()
                        cat = clasificar_tecnologia_yupana(c_clean, arch if arch in ["HIDRO", "RER"] else "TERMICA")
                        
                        match = False
                        if tipo_principal == "TERMICA" and cat in ["Biogás+Biomasa+Nafta+Flexigas", "Gas de Camisea", "Gas del Norte+Gas de la Selva", "Residual+Diésel D2"]: match = True
                        elif tipo_principal == "HIDRO" and cat == "Hidráulica": match = True
                        elif tipo_principal == "SOLAR" and cat == "Solar": match = True
                        elif tipo_principal == "EOLICO" and cat == "Eólica": match = True
                        
                        if match:
                            if k not in dia_data: dia_data[k] = [0.0]*48
                            dia_data[k][i] += rellenar_hasta_48(v_list)[i]
                        
            df_dia = pd.DataFrame(dia_data)
            df_dia.insert(0, 'Hora', ts_dict[f])
            dfs_tab.append(df_dia)
            
        if dfs_tab:
            df_total = pd.concat(dfs_tab, ignore_index=True).fillna(0)
            num_cols = [c for c in df_total.columns if c != 'Hora']
            active_cols = [c for c in num_cols if df_total[c].sum() > 0]
            
            if not active_cols:
                st.warning("No hay generación para este tipo en las fechas seleccionadas.")
                return
                
            todas_centrales = sorted(active_cols)
            filtro = st.multiselect(f"⚡ Filtrar Centrales:", options=todas_centrales, default=[], placeholder="Todas (vacío) o buscar...")
            lista_filtro = filtro if filtro else todas_centrales
            
            df_plot = df_total[['Hora'] + lista_filtro]
            st.plotly_chart(crear_grafica_area_apilada(df_plot, marcadores=marcadores_globales), use_container_width=True)

    with t_hidro:
        st.markdown("### 💧 Despacho Hidroeléctrico Continuo")
        render_tab_generico("HIDRO")

    with t_term:
        st.markdown("### 🔥 Despacho Térmico Continuo")
        st.info("Consolida el **Despacho Activo Real** de unidades Térmicas y Biomasa. Totalmente homologado a la Matriz de Generación Total.")
        render_tab_generico("TERMICA")

    with t_eol:
        st.markdown("### 💨 Despacho Eólico Continuo")
        render_tab_generico("EOLICO")

    with t_sol:
        st.markdown("### ☀️ Despacho Solar Continuo")
        render_tab_generico("SOLAR")

    # === INACTIVA DIÉSEL CONTINUA (CÁLCULO POR INTERVALO) ===
    with t_inactiva:
        st.markdown("### 🛑 Capacidad Inactiva Diésel/Residual")
        st.info("Mapea la capacidad disponible no despachada en cada intervalo de media hora. Si la central enciende en un intervalo, su gráfica cae a 0 MW en esa hora. Ciclos combinados están excluidos.")
        
        dfs_inactiva = []
        mantenimiento_global = []
        
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            diesel_plants = set()
            for p in set(active_prog):
                for arch in ["TERMICA", "POT_EFEC"]:
                    for c in dics_cache[p][arch].keys():
                        if clasificar_tecnologia_yupana(c, arch) == "Residual+Diésel D2":
                            diesel_plants.add(c)
                            
            inactiva_dia = {}
            for c in diesel_plants:
                efec_day = [0.0]*48
                desp_day = [0.0]*48
                for i in range(48):
                    p = active_prog[i]
                    if c in dics_cache[p]["TERMICA"]: desp_day[i] += rellenar_hasta_48(dics_cache[p]["TERMICA"][c])[i]
                    if c in dics_cache[p]["RER"]: desp_day[i] += rellenar_hasta_48(dics_cache[p]["RER"][c])[i]
                    if c in dics_cache[p]["POT_EFEC"]: efec_day[i] += rellenar_hasta_48(dics_cache[p]["POT_EFEC"][c])[i]
                    
                sum_efec = sum(efec_day)
                
                if sum_efec == 0:
                    mantenimiento_global.append({"Fecha": f.strftime('%d/%m/%Y'), "Central": c, "Tecnología": "Residual+Diésel D2", "Estado Operativo": "Mantenimiento / Fuera de Servicio"})
                else:
                    # Reserva Inactiva: Si opera (>0) en la hora, se excluye totalmente de la inactiva en ese intervalo
                    idle = [0.0 if desp_day[i] > 0 else max(0.0, round(efec_day[i], 2)) for i in range(48)]
                    if sum(idle) > 0:
                        inactiva_dia[c] = idle
                        
            df_dia = pd.DataFrame(inactiva_dia) if inactiva_dia else pd.DataFrame()
            df_dia['Hora'] = ts_dict[f]
            dfs_inactiva.append(df_dia)
            
        if dfs_inactiva:
            df_total_inac = pd.concat(dfs_inactiva, ignore_index=True).fillna(0)
            num_cols = [c for c in df_total_inac.columns if c != 'Hora']
            
            if sum(df_total_inac[num_cols].sum()) > 0:
                st.markdown("#### 📉 Capacidad Inactiva Detallada")
                st.plotly_chart(crear_grafica_area_apilada(df_total_inac, marcadores=marcadores_globales, aplicar_colores=False), use_container_width=True)
                
                st.markdown("#### 📊 Capacidad Inactiva Acumulada Total")
                df_acum = pd.DataFrame({"Hora": df_total_inac['Hora'], "Total Diésel Inactivo": df_total_inac[num_cols].sum(axis=1)})
                fig_acum = px.area(df_acum, x="Hora", y="Total Diésel Inactivo", color_discrete_sequence=["#FF0000"])
                
                for trace in fig_acum.data:
                    trace.hovertemplate = '<b>%{y:,.2f} MW</b><br>%{x|%d/%m %H:%M}'
                    trace.hoverinfo = ['skip' if pd.isna(y) or y <= 0.01 else 'all' for y in trace.y]
                
                for ts, txt in marcadores_globales:
                    fig_acum.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(255,255,255,0.7)")
                    txt_limpio = txt.replace("(", "").replace(")", "")
                    txt_final = f"{txt_limpio} {ts.strftime('%H:%M')}"
                    align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                    fig_acum.add_annotation(x=ts, y=1.02, yref="paper", text=f"<b>{txt_final}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", bordercolor="white", borderwidth=1, borderpad=3, textangle=-90, yanchor="bottom", xanchor=align)

                fig_acum.update_layout(hovermode="x unified", height=550, margin=dict(t=150, b=50, l=60, r=50))
                st.plotly_chart(fig_acum, use_container_width=True)
            else:
                st.success("✅ Toda la capacidad Diésel/Residual disponible fue despachada o no hubo potencia inactiva en el periodo.")
                
        if mantenimiento_global:
            st.markdown("#### 🛠️ Centrales en Mantenimiento (0 MW Efectivo)")
            st.dataframe(pd.DataFrame(mantenimiento_global), use_container_width=True)

    # === DEMANDA Y MATRIZ ENERGÉTICA (DESPACHO ACTIVO REAL) ===
    with t_dem:
        st.markdown("### 📈 Demanda Total del Sistema")
        dfs_demanda = []
        dfs_matriz = []
        
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            cats = {k: [0.0]*48 for k in COLOR_MAP.keys()}
            tot_sis = [0.0]*48
            
            for i in range(48):
                p = active_prog[i]
                for arch in ["HIDRO", "TERMICA", "RER"]:
                    d = dics_cache[p][arch]
                    for k, v_list in d.items():
                        cat = clasificar_tecnologia_yupana(k, arch)
                        val = rellenar_hasta_48(v_list)[i]
                        cats[cat][i] += val
                        tot_sis[i] += val
                        
            df_dem_dia = pd.DataFrame({"Demanda Efectiva (MW)": tot_sis})
            df_dem_dia.insert(0, 'Hora', ts_dict[f])
            dfs_demanda.append(df_dem_dia)
            
            df_mat_dia = pd.DataFrame(cats)
            df_mat_dia.insert(0, 'Hora', ts_dict[f])
            dfs_matriz.append(df_mat_dia)
            
        if dfs_demanda:
            df_dem_total = pd.concat(dfs_demanda, ignore_index=True)
            fig_dem = px.line(df_dem_total, x="Hora", y="Demanda Efectiva (MW)", markers=True)
            
            for trace in fig_dem.data:
                trace.hovertemplate = '<b>%{y:,.2f} MW</b><br>%{x|%d/%m %H:%M}'
            
            for ts, txt in marcadores_globales:
                fig_dem.add_vline(x=ts, line_width=1, line_dash="dash", line_color="grey")
                txt_limpio = txt.replace("(", "").replace(")", "")
                txt_final = f"{txt_limpio} {ts.strftime('%H:%M')}"
                align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                fig_dem.add_annotation(x=ts, y=1.02, yref="paper", text=f"<b>{txt_final}</b>", showarrow=False, font=dict(size=10, color="black"), bgcolor="lightgrey", textangle=-90, yanchor="bottom", xanchor=align)
                
            fig_dem.update_layout(hovermode="x unified", height=500, margin=dict(t=150, l=60))
            st.plotly_chart(fig_dem, use_container_width=True)

        st.markdown("---")
        st.markdown("### 📊 Matriz de Generación Acumulada por Tecnología")
        st.info("Distribución energética global en base al Despacho Activo Real. Respeta al 100% las categorizaciones de combustible y no suma reservas rodantes.")
        
        if dfs_matriz:
            df_mat_total = pd.concat(dfs_matriz, ignore_index=True)
            num_cols = df_mat_total.columns.drop('Hora')
            df_mat_total = df_mat_total.loc[:, (df_mat_total != 0).any(axis=0) | (df_mat_total.columns == 'Hora')]
            
            st.plotly_chart(crear_grafica_area_apilada(
                df_mat_total, 
                marcadores=marcadores_globales, aplicar_colores=True, orden_fijo=ORDEN_TECNOLOGIAS
            ), use_container_width=True)

        st.markdown("---")
        st.markdown("### 🗄️ Datos Fuente: Despacho Continuo por Central")
        
        active_prog_global = []
        for f in fechas_ordenadas:
            if f in active_prog_dict:
                active_prog_global.extend(active_prog_dict[f])
        
        df_stitched_raw = pd.DataFrame({'Hora': [ts.strftime("%Y-%m-%d %H:%M") for ts in timestamps_globales], "Programa_Vigente": active_prog_global})
        
        todas_las_centrales = {"HIDRO": set(), "TERMICA": set(), "RER": set()}
        for f in fechas_ordenadas:
            if f in dics_cache_dict:
                for p in set(active_prog_dict[f]):
                    todas_las_centrales["HIDRO"].update(dics_cache_dict[f][p]["HIDRO"].keys())
                    todas_las_centrales["TERMICA"].update(dics_cache_dict[f][p]["TERMICA"].keys())
                    todas_las_centrales["RER"].update(dics_cache_dict[f][p]["RER"].keys())

        todas_activas = set(todas_las_centrales["HIDRO"].union(todas_las_centrales["TERMICA"], todas_las_centrales["RER"]))
        
        dic_export = {}
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            for i in range(48):
                p = active_prog[i]
                for c in todas_activas:
                    tipo_origen = "HIDRO" if c in todas_las_centrales["HIDRO"] else "TERMICA" if c in todas_las_centrales["TERMICA"] else "RER"
                    cat = clasificar_tecnologia_yupana(c, tipo_origen)
                    nombre_columna = f"{c} ({cat})"
                    
                    if nombre_columna not in dic_export: dic_export[nombre_columna] = []
                    
                    val = 0.0
                    for arch in ["HIDRO", "TERMICA", "RER"]:
                        if c in dics_cache[p][arch]:
                            val += rellenar_hasta_48(dics_cache[p][arch][c])[i]
                            
                    dic_export[nombre_columna].append(round(val, 2))
                    
        for k, v_list in dic_export.items():
            if sum(v_list) > 0:
                df_stitched_raw[k] = v_list
                
        st.dataframe(df_stitched_raw, use_container_width=True)
        
        buffer_raw = io.BytesIO()
        with pd.ExcelWriter(buffer_raw, engine='openpyxl') as writer:
            df_stitched_raw.to_excel(writer, index=False, sheet_name='Despacho_Continuo')
        st.download_button(
            label="📥 Descargar Matriz Empalmada Completa (Excel)", data=buffer_raw.getvalue(),
            file_name=f"Despacho_MultiDia_SEIN_{ini.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary"
        )

    # === MOTIVOS DE REPROGRAMAS ===
    with t_motivos_rdo:
        st.markdown("### 📋 Motivos de Reprogramación Operativa")
        st.info("Extraído directamente de la columna C del archivo Excel original de cada reprograma subido por el COES.")
        
        tabla_motivos = []
        for f in fechas_ordenadas:
            if f not in dics_cache_dict: continue
            progs = active_prog_dict[f]
            
            for p in sorted(set(progs)):
                if "RDO" in p:
                    motivo_texto = data[f]["Dataframes"].get(f"MOTIVO_{p}", "Motivo no disponible en el sistema.")
                    tabla_motivos.append({
                        "Fecha": f.strftime("%d/%m/%Y"),
                        "Reprograma": p,
                        "Justificación / Motivo": motivo_texto
                    })
                    
        if tabla_motivos:
            st.dataframe(pd.DataFrame(tabla_motivos), use_container_width=True)
        else:
            st.success("No hay reprogramas en el rango seleccionado, o no se encontró justificación en los archivos de origen.")